import openpyxl

# Como eu crio uma planilha? (book)
book = openpyxl.Workbook()
# Como visualizat  páginas existentes? 
print(book.sheetnames)
# Como criar eu crio uma página para um banco de dados para uma empresa por exemplo?
book.create_sheet('Frutas')
# Como selecionar uma página:
frutas_page = book['Frutas']
frutas_page.append(['Banana','5','R$3,90'])
frutas_page.append(['Limao','3','R$2,50'])
frutas_page.append(['Laranja','2','R$1,60'])
frutas_page.append(['Pera','6','R$1,10'])
# logo eu irei salvar a planilha
book.save('Planilha de Compras.xlsx')