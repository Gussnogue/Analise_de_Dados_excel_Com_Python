import openpyxl

# Carregando arquivo
book = openpyxl.load_workbook('Planilha de Compras.xlsx')
# Selecionando uma página
frutas_page = book['Frutas']
# Imprimido os dados de cada linha
for rows in frutas_page.iter_rows(min_row=1, max_row=4):
    for cell in rows:
        print(cell.value)

        #para sair organizado só trocar por:
#for rows in frutas_page.iter_rows(min_row=1, max_row=4):
    #print(rows[0].value,rows[1].value,rows[2].value)